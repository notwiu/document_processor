import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

class ExcelHandler:
    def __init__(self):
        pass
    
    def save_tables_to_excel(self, tables, output_path):
        """Salva múltiplas tabelas em um arquivo Excel"""
        if not tables:
            return False
        
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for i, table_info in enumerate(tables):
                    df = table_info['dataframe']
                    sheet_name = f"Pag{table_info['page']}_Tab{i+1}"
                    
                    # Limitar tamanho do nome da sheet
                    if len(sheet_name) > 31:
                        sheet_name = sheet_name[:31]
                    
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    # Formatar a planilha
                    self._format_worksheet(writer.book[sheet_name], df)
            
            return True
            
        except Exception as e:
            print(f"Erro ao salvar Excel: {e}")
            return False
    
    def _format_worksheet(self, ws, df):
        """Formata a planilha do Excel"""
        # Definir estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Formatar cabeçalho
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border
        
        # Formatar células de dados
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                cell.border = thin_border
                # Alinhar números à direita, texto à esquerda
                try:
                    float(cell.value)
                    cell.alignment = Alignment(horizontal="right")
                except (ValueError, TypeError):
                    cell.alignment = Alignment(horizontal="left")
        
        # Ajustar largura das colunas
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def create_summary_report(self, tables, output_path):
        """Cria um relatório resumido das tabelas extraídas"""
        summary_data = []
        
        for table_info in tables:
            summary_data.append({
                'Página': table_info['page'],
                'Tabela': table_info['table_number'],
                'Linhas': table_info['shape'][0],
                'Colunas': table_info['shape'][1],
                'Colunas (nomes)': ', '.join(str(col) for col in table_info['columns'][:3]) + 
                                   ('...' if len(table_info['columns']) > 3 else '')
            })
        
        if summary_data:
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(output_path, index=False)
            return True
        
        return False